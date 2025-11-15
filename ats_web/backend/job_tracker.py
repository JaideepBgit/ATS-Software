"""
Job Application Tracker for Web ATS
Manages job applications and updates Excel tracking sheet
"""
import os
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List
import openpyxl
from openpyxl import Workbook, load_workbook


class JobTracker:
    """Track job applications in Excel"""
    
    def __init__(self, excel_path: str = "data/jobs_applied/job_applicaiton.xlsx"):
        self.excel_path = Path(excel_path)
        self._ensure_excel_exists()
    
    def _ensure_excel_exists(self):
        """Create Excel file if it doesn't exist"""
        if not self.excel_path.exists():
            # Create directory if needed
            self.excel_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create new workbook with headers
            wb = Workbook()
            ws = wb.active
            ws.title = "Jobs Applied"
            
            # Set headers
            headers = ["Company", "Job", "Portal", "Full Time", "Date Applied"]
            ws.append(headers)
            
            # Save
            wb.save(self.excel_path)
            print(f"✓ Created new job tracking Excel: {self.excel_path}")
    
    def add_job_application(self, company: str, job_title: str, 
                           portal: str = "LinkedIn", 
                           employment_type: str = "Full Time") -> Dict:
        """Add a new job application to the Excel sheet
        
        Args:
            company: Company name
            job_title: Job title/position
            portal: Job portal (default: LinkedIn)
            employment_type: Employment type (default: Full Time)
            
        Returns:
            Dict with success status and details
        """
        try:
            # Load workbook
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Add new row
            date_applied = datetime.now().strftime("%Y-%m-%d")
            ws.append([company, job_title, portal, employment_type, date_applied])
            
            # Save
            wb.save(self.excel_path)
            
            return {
                "success": True,
                "company": company,
                "job_title": job_title,
                "portal": portal,
                "employment_type": employment_type,
                "date_applied": date_applied,
                "total_applications": self.get_application_count()
            }
            
        except Exception as e:
            print(f"Error logging job application: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def get_application_count(self) -> int:
        """Get total number of applications"""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            # Subtract 1 for header row
            return ws.max_row - 1
        except:
            return 0
    
    def check_if_applied(self, company: str, job_title: str) -> bool:
        """Check if already applied to this job
        
        Args:
            company: Company name
            job_title: Job title
            
        Returns:
            True if already applied, False otherwise
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Check each row (skip header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Company and Job columns
                    if (row[0].lower().strip() == company.lower().strip() and 
                        row[1].lower().strip() == job_title.lower().strip()):
                        return True
            
            return False
            
        except Exception as e:
            print(f"Error checking applications: {str(e)}")
            return False
    
    def get_recent_applications(self, limit: int = 10) -> List[Dict]:
        """Get recent job applications
        
        Args:
            limit: Number of recent applications to return
            
        Returns:
            List of recent applications
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            applications = []
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            # Get last N rows
            for row in rows[-limit:]:
                if row[0]:  # Has company name
                    applications.append({
                        'company': row[0],
                        'job': row[1],
                        'portal': row[2],
                        'type': row[3],
                        'date': row[4]
                    })
            
            return list(reversed(applications))  # Most recent first
            
        except Exception as e:
            print(f"Error getting recent applications: {str(e)}")
            return []
    
    def get_all_applications(self) -> List[Dict]:
        """Get all job applications
        
        Returns:
            List of all applications
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            applications = []
            
            # Get all rows (skip header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Has company name
                    applications.append({
                        'company': row[0],
                        'job': row[1],
                        'portal': row[2],
                        'type': row[3],
                        'date': row[4]
                    })
            
            return list(reversed(applications))  # Most recent first
            
        except Exception as e:
            print(f"Error getting all applications: {str(e)}")
            return []
    
    def get_statistics(self) -> Dict:
        """Get application statistics
        
        Returns:
            Dict with statistics
        """
        try:
            applications = self.get_all_applications()
            
            if not applications:
                return {
                    "total": 0,
                    "by_portal": {},
                    "by_type": {},
                    "recent_7_days": 0
                }
            
            # Count by portal
            by_portal = {}
            for app in applications:
                portal = app.get('portal', 'Unknown')
                by_portal[portal] = by_portal.get(portal, 0) + 1
            
            # Count by type
            by_type = {}
            for app in applications:
                emp_type = app.get('type', 'Unknown')
                by_type[emp_type] = by_type.get(emp_type, 0) + 1
            
            # Count recent (last 7 days)
            from datetime import datetime, timedelta
            seven_days_ago = datetime.now() - timedelta(days=7)
            recent_count = 0
            
            for app in applications:
                try:
                    app_date = datetime.strptime(str(app.get('date', '')), "%Y-%m-%d")
                    if app_date >= seven_days_ago:
                        recent_count += 1
                except:
                    pass
            
            return {
                "total": len(applications),
                "by_portal": by_portal,
                "by_type": by_type,
                "recent_7_days": recent_count
            }
            
        except Exception as e:
            print(f"Error getting statistics: {str(e)}")
            return {
                "total": 0,
                "by_portal": {},
                "by_type": {},
                "recent_7_days": 0
            }
