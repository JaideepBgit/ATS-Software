"""
Job Application Tracker
Manages job applications and updates Excel tracking sheet
"""
import os
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict
import openpyxl
from openpyxl import Workbook, load_workbook


class JobTracker:
    """Track job applications in Excel"""
    
    def __init__(self, excel_path: str = "data/jobs_applied/job_applicaiton.xlsx"):
        self.excel_path = Path(excel_path)
        self._ensure_excel_exists()
    
    def _ensure_excel_exists(self):
        """Create Excel file if it doesn't exist"""
        if not self.excel_path.exists():
            # Create directory if needed
            self.excel_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create new workbook with headers
            wb = Workbook()
            ws = wb.active
            ws.title = "Jobs Applied"
            
            # Set headers
            headers = ["Company", "Job", "Portal", "Full Time", "Date Applied"]
            ws.append(headers)
            
            # Save
            wb.save(self.excel_path)
            print(f"✓ Created new job tracking Excel: {self.excel_path}")
    
    def add_job_application(self, company: str, job_title: str, 
                           portal: str = "LinkedIn", 
                           employment_type: str = "Full Time") -> bool:
        """Add a new job application to the Excel sheet
        
        Args:
            company: Company name
            job_title: Job title/position
            portal: Job portal (default: LinkedIn)
            employment_type: Employment type (default: Full Time)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Load workbook
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Add new row
            date_applied = datetime.now().strftime("%Y-%m-%d")
            ws.append([company, job_title, portal, employment_type, date_applied])
            
            # Save
            wb.save(self.excel_path)
            
            print(f"\n✅ Job application logged!")
            print(f"   Company: {company}")
            print(f"   Position: {job_title}")
            print(f"   Portal: {portal}")
            print(f"   Type: {employment_type}")
            print(f"   Date: {date_applied}")
            
            return True
            
        except Exception as e:
            print(f"\n❌ Error logging job application: {str(e)}")
            return False
    
    def get_application_count(self) -> int:
        """Get total number of applications"""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            # Subtract 1 for header row
            return ws.max_row - 1
        except:
            return 0
    
    def check_if_applied(self, company: str, job_title: str) -> bool:
        """Check if already applied to this job
        
        Args:
            company: Company name
            job_title: Job title
            
        Returns:
            True if already applied, False otherwise
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Check each row (skip header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Company and Job columns
                    if (row[0].lower().strip() == company.lower().strip() and 
                        row[1].lower().strip() == job_title.lower().strip()):
                        return True
            
            return False
            
        except Exception as e:
            print(f"⚠️  Error checking applications: {str(e)}")
            return False
    
    def get_recent_applications(self, limit: int = 5) -> list:
        """Get recent job applications
        
        Args:
            limit: Number of recent applications to return
            
        Returns:
            List of recent applications
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            applications = []
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            # Get last N rows
            for row in rows[-limit:]:
                if row[0]:  # Has company name
                    applications.append({
                        'company': row[0],
                        'job': row[1],
                        'portal': row[2],
                        'type': row[3],
                        'date': row[4]
                    })
            
            return list(reversed(applications))  # Most recent first
            
        except Exception as e:
            print(f"⚠️  Error getting recent applications: {str(e)}")
            return []
